"""
Enrich Forbes Global 2000 with business descriptions
=====================================================
For each of the ~2,000 Forbes companies, finds the matching Yahoo Finance
ticker via name search and fetches the longBusinessSummary description.

Why this matters
----------------
Within broad Forbes industry categories the P/S multiple spreads enormously:
  IT Software & Services : 0.1× (Ingram Micro) → 586× (Knowledge Atlas)
  Transportation         : 0.08× (Air France)  →  12.6× (Transurban toll roads)
  Utilities              : 0.1× (Tokyo Electric) → 178× (Power Assets Holdings)

A company description lets us distinguish "SaaS / AI" from "IT distribution",
"airport concession" from "airline", "renewable hydro" from "fossil fuel utility".
That makes comparable-company matching far more accurate.

Output
------
  data/forbes_enriched.pkl   — dict with full DataFrame + metadata
  data/forbes_enriched.xlsx  — same data, human-readable

Usage
-----
  python scripts/enrich_forbes_descriptions.py

Expected runtime: 10–20 minutes (2,000 companies, 12 concurrent workers).
Expected coverage: ~65–75 % of Forbes companies will get a description
(misses: private companies, small regional firms, Chinese companies whose
English name doesn't resolve cleanly in Yahoo Finance).
"""

import pickle
import re
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf

sys.path.insert(0, str(Path(__file__).parent))
from fetch_yf_training import get_usd_rate  # noqa: E402 — thread-safe, cached FX lookup

INPUT   = Path(__file__).parent.parent / "data" / "forbes_global2000.xlsx"
OUT_PKL = Path(__file__).parent.parent / "data" / "forbes_enriched.pkl"
OUT_XLS = Path(__file__).parent.parent / "data" / "forbes_enriched.xlsx"

# ---------------------------------------------------------------------------
# Name-matching helpers (same logic used in app.py)
# ---------------------------------------------------------------------------

_SUFFIXES = [
    " incorporated", " corporation", " holdings", " holding",
    " limited", " group", " company", " plc", " inc.", " inc",
    " corp.", " corp", " ltd.", " ltd", " llc", " ag", " nv",
    " sa", " s.a.", " co.", " co", " a/s", " asa", " ab",
    " technologies", " technology", " systems", " solutions",
    " international", " enterprises", " services", " industries",
]

def _clean(name: str) -> str:
    n = name.lower().strip()
    n = re.sub(r"[&,\.\-']", " ", n)
    for s in _SUFFIXES:
        n = n.replace(s, "")
    return re.sub(r"\s+", " ", n).strip()


def _name_score(forbes_name: str, yf_name: str) -> float:
    """0–1 score for how well the YF result name matches the Forbes name."""
    a = _clean(forbes_name)
    b = _clean(yf_name)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if len(shorter) >= 4 and bool(re.search(r"\b" + re.escape(shorter), longer)):
        return 0.85
    # Partial word overlap (Jaccard on tokens)
    ta = set(a.split())
    tb = set(b.split())
    if not ta or not tb:
        return 0.0
    overlap = len(ta & tb) / len(ta | tb)
    return overlap


# Country → ISO-2 hints for exchange suffix matching
_EXCHANGE_HINTS: dict[str, list[str]] = {
    "Japan":          [".T"],
    "South Korea":    [".KS", ".KQ"],
    "China":          [".SS", ".SZ", ".HK"],
    "Hong Kong":      [".HK"],
    "Taiwan":         [".TW"],
    "India":          [".NS", ".BO"],
    "Australia":      [".AX"],
    "United Kingdom": [".L"],
    "Germany":        [".DE", ".F"],
    "France":         [".PA"],
    "Italy":          [".MI"],
    "Spain":          [".MC"],
    "Denmark":        [".CO"],
    "Norway":         [".OL"],
    "Sweden":         [".ST"],
    "Switzerland":    [".SW"],
    "Netherlands":    [".AS"],
    "Saudi Arabia":   [".SR"],
    "Singapore":      [".SI"],
    "Brazil":         [".SA"],
    "Canada":         [".TO", ".V"],
}


def _country_score(forbes_country: str, yf_country: str, ticker: str) -> float:
    if not yf_country or not forbes_country:
        return 0.0
    if forbes_country.lower() == yf_country.lower():
        return 1.0
    # Check exchange suffix hints
    hints = _EXCHANGE_HINTS.get(forbes_country, [])
    if any(ticker.upper().endswith(h) for h in hints):
        return 0.8
    return 0.0


# ---------------------------------------------------------------------------
# Per-company fetcher
# ---------------------------------------------------------------------------

_print_lock = threading.Lock()

def _extract_financials(info: dict) -> dict:
    """Pull standardised financial metrics from a yfinance info dict."""
    fx = get_usd_rate(info.get("financialCurrency") or info.get("currency") or "USD")
    total_debt = (info.get("totalDebt") or 0) * fx
    total_cash = ((info.get("totalCash")
                  or info.get("totalCashAndShortTermInvestments") or 0)) * fx
    net_debt_B = (total_debt - total_cash) / 1e9
    fcf = info.get("freeCashflow")
    tr  = info.get("totalRevenue")
    fcf_margin = round(fcf / tr, 4) if fcf and tr else None
    return {
        "net_debt_B":    round(net_debt_B, 3),
        "rev_growth":    info.get("revenueGrowth"),
        "ebitda_margin": info.get("ebitdaMargins"),
        "gross_margin":  info.get("grossMargins"),
        "fcf_margin":    fcf_margin,
    }


def fetch_description(row: dict) -> dict:
    """
    Given a Forbes row, search Yahoo Finance for the best matching ticker
    and return {ticker, description, net_debt_B, yf_name, match_score,
    rev_growth, ebitda_margin, gross_margin, fcf_margin}.
    Returns a dict with empty strings if no good match found.
    """
    name    = str(row["Company"]).strip()
    country = str(row.get("Country", "")).strip()
    result  = {"ticker": "", "description": "", "yf_name": "", "match_score": 0.0,
               "net_debt_B": 0.0, "rev_growth": None, "ebitda_margin": None,
               "gross_margin": None, "fcf_margin": None}

    try:
        # Search Yahoo Finance by company name
        search_results = yf.Search(name, max_results=6).quotes
        if not search_results:
            return result

        best_score  = 0.0
        best_symbol = None
        best_yf_name = ""

        for r in search_results:
            if r.get("quoteType") not in ("EQUITY", "MUTUALFUND"):
                # Prefer EQUITYs; skip crypto, ETFs, etc.
                if r.get("quoteType") != "EQUITY":
                    continue
            symbol   = r.get("symbol", "")
            yf_name  = r.get("longname") or r.get("shortname", "")
            yf_ctry  = r.get("country", "")

            ns = _name_score(name, yf_name)
            cs = _country_score(country, yf_ctry, symbol)
            # Combined score: name match is the primary signal
            score = ns * 0.7 + cs * 0.3

            if score > best_score:
                best_score  = score
                best_symbol = symbol
                best_yf_name = yf_name

        if best_symbol is None or best_score < 0.25:
            return result   # no acceptable match

        # Fetch description + financials from best-match ticker
        info = yf.Ticker(best_symbol).info
        desc = info.get("longBusinessSummary", "")

        result["ticker"]      = best_symbol
        result["description"] = desc
        result["yf_name"]     = best_yf_name
        result["match_score"] = round(best_score, 3)
        result.update(_extract_financials(info))
        return result

    except Exception:
        return result


def fetch_extra_fields(ticker: str) -> dict:
    """
    Fetch financial metrics for a company whose ticker is already known.
    Uses Ticker.info directly — no Search API, so no 401 rate-limit risk.
    """
    try:
        info = yf.Ticker(ticker).info
        if not info.get("shortName"):
            return {}
        return _extract_financials(info)
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def enrich(df: pd.DataFrame, max_workers: int = 12) -> pd.DataFrame:
    total = len(df)
    rows  = df.to_dict("records")
    results: list[dict] = [{}] * total
    done  = 0

    print(f"\nEnriching {total} companies with {max_workers} workers …\n")
    t0 = time.time()

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = {ex.submit(fetch_description, row): i for i, row in enumerate(rows)}
        for fut in as_completed(futs):
            idx = futs[fut]
            try:
                res = fut.result()
            except Exception:
                res = {"ticker": "", "description": "", "yf_name": "", "match_score": 0.0}
            results[idx] = res
            done += 1

            if done % 50 == 0 or done == total:
                elapsed = time.time() - t0
                pct     = done / total * 100
                with_desc = sum(1 for r in results[:done] if r.get("description"))
                eta = elapsed / done * (total - done)
                with _print_lock:
                    print(f"  [{done:4d}/{total}] {pct:5.1f}%  "
                          f"descriptions: {with_desc}  "
                          f"elapsed: {elapsed/60:.1f}m  "
                          f"ETA: {eta/60:.1f}m")

    out = df.copy()
    out["ticker"]       = [r.get("ticker", "")        for r in results]
    out["yf_name"]      = [r.get("yf_name", "")       for r in results]
    out["description"]  = [r.get("description", "")   for r in results]
    out["net_debt_B"]   = [r.get("net_debt_B", 0.0)   for r in results]
    out["match_score"]  = [r.get("match_score", 0.0)  for r in results]
    out["rev_growth"]   = [r.get("rev_growth")         for r in results]
    out["ebitda_margin"]= [r.get("ebitda_margin")      for r in results]
    out["gross_margin"] = [r.get("gross_margin")       for r in results]
    out["fcf_margin"]   = [r.get("fcf_margin")         for r in results]
    return out


def recompute_net_debt(df: pd.DataFrame, max_workers: int = 12) -> pd.DataFrame:
    """
    Recompute net_debt_B for every company with a resolved ticker, using the
    now-fx-aware _extract_financials(). Unlike augment_financials(), this
    always overwrites net_debt_B — the existing stored value is wrong for
    non-USD companies (pre-fix), not missing, so a null-check wouldn't select
    it. Uses Ticker.info directly, same as augment_financials() — no Search
    API calls, so no 401 crumb-expiry rate-limit risk.
    """
    needs = df[df["ticker"].str.len() > 0]
    if needs.empty:
        print("  Recompute: no companies with tickers to correct.")
        return df

    tickers = needs["ticker"].tolist()
    indices = needs.index.tolist()
    print(f"\nRecomputing net_debt_B for {len(tickers)} companies with known tickers …")

    t0 = time.time()
    results: dict[int, dict] = {}
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = {ex.submit(fetch_extra_fields, tk): idx
                for tk, idx in zip(tickers, indices)}
        done = 0
        for fut in as_completed(futs):
            idx = futs[fut]
            try:
                res = fut.result()
            except Exception:
                res = {}
            results[idx] = res
            done += 1
            if done % 100 == 0 or done == len(tickers):
                elapsed = time.time() - t0
                print(f"  [{done:4d}/{len(tickers)}]  elapsed: {elapsed/60:.1f}m")

    corrected = 0
    for idx, fields in results.items():
        if "net_debt_B" in fields and fields["net_debt_B"] is not None:
            df.at[idx, "net_debt_B"] = fields["net_debt_B"]
            corrected += 1

    print(f"  Recompute complete: {corrected}/{len(tickers)} companies had net_debt_B corrected.")
    return df


def fix_net_debt_currency() -> None:
    """
    One-time corrective pass: recompute net_debt_B for every already-ticketed
    company in the existing forbes_enriched.pkl, using the fx-aware
    _extract_financials(). Backs up the current pkl/xlsx first (no version
    control in this workspace, so this is the only rollback path). Does not
    touch descriptions, tickers, or other financial metrics — see
    recompute_net_debt().
    """
    if not OUT_PKL.exists():
        print(f"No existing pickle at {OUT_PKL} — nothing to correct. "
              f"Run without --fix-net-debt first to build the initial dataset.")
        return

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_pkl = OUT_PKL.with_name(f"{OUT_PKL.stem}.backup-{stamp}{OUT_PKL.suffix}")
    backup_xls = OUT_XLS.with_name(f"{OUT_XLS.stem}.backup-{stamp}{OUT_XLS.suffix}")
    backup_pkl.write_bytes(OUT_PKL.read_bytes())
    if OUT_XLS.exists():
        backup_xls.write_bytes(OUT_XLS.read_bytes())
    print(f"Backed up existing data to {backup_pkl.name} (and .xlsx if present).")

    with open(OUT_PKL, "rb") as f:
        prior = pickle.load(f)
    df = prior["df"]
    n_desc_before = (df["description"].str.len() > 50).sum()

    df = recompute_net_debt(df, max_workers=12)

    n_desc_after = (df["description"].str.len() > 50).sum()
    if n_desc_after != n_desc_before:
        print(f"  WARNING: description coverage changed ({n_desc_before} -> {n_desc_after}) "
              f"— this pass should only touch net_debt_B.")

    save(df)


def augment_financials(df: pd.DataFrame, max_workers: int = 12) -> pd.DataFrame:
    """
    For companies that already have a ticker but are missing financial metrics
    (rev_growth, ebitda_margin, gross_margin, fcf_margin), fetch them directly
    via Ticker.info — no Search API, so no 401 rate-limit risk.
    """
    METRIC_COLS = ["rev_growth", "ebitda_margin", "gross_margin", "fcf_margin"]
    for col in METRIC_COLS:
        if col not in df.columns:
            df[col] = None

    # Only augment rows with a ticker but at least one missing metric
    needs = df[
        (df["ticker"].str.len() > 0) &
        df[METRIC_COLS].isnull().any(axis=1)
    ]
    if needs.empty:
        print("  Augment: all companies with tickers already have financial metrics.")
        return df

    tickers = needs["ticker"].tolist()
    indices = needs.index.tolist()
    print(f"\nAugmenting financial metrics for {len(tickers)} companies with known tickers …")

    t0 = time.time()
    results: dict[int, dict] = {}
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = {ex.submit(fetch_extra_fields, tk): idx
                for tk, idx in zip(tickers, indices)}
        done = 0
        for fut in as_completed(futs):
            idx = futs[fut]
            try:
                res = fut.result()
            except Exception:
                res = {}
            results[idx] = res
            done += 1
            if done % 100 == 0 or done == len(tickers):
                elapsed = time.time() - t0
                print(f"  [{done:4d}/{len(tickers)}]  elapsed: {elapsed/60:.1f}m")

    for idx, fields in results.items():
        for col in METRIC_COLS:
            if col in fields and fields[col] is not None:
                df.at[idx, col] = fields[col]

    filled = (df.loc[indices, METRIC_COLS].notnull().any(axis=1)).sum()
    print(f"  Augment complete: {filled}/{len(tickers)} companies got at least one new metric.")
    return df


def save(df: pd.DataFrame) -> None:
    n_desc  = (df["description"].str.len() > 50).sum()
    n_total = len(df)
    coverage = n_desc / n_total * 100

    # ── Pickle ──────────────────────────────────────────────────────────────
    meta = {
        "df":          df,
        "fetched_at":  datetime.now().isoformat(),
        "n_companies": n_total,
        "n_with_desc": n_desc,
        "coverage_pct": round(coverage, 1),
    }
    OUT_PKL.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PKL, "wb") as f:
        pickle.dump(meta, f)
    print(f"\nSaved pickle  → {OUT_PKL}")

    # ── Excel ────────────────────────────────────────────────────────────────
    with pd.ExcelWriter(OUT_XLS, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Forbes Enriched")
        ws = writer.sheets["Forbes Enriched"]
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # Column widths
        col_widths = {
            "A": 7,  "B": 35, "C": 20, "D": 30, "E": 25,
            "F": 10, "G": 10, "H": 10, "I": 12,
            "J": 15, "K": 15, "L": 30, "M": 70,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
        ws.freeze_panes = "A2"
    print(f"Saved Excel   → {OUT_XLS}")

    # ── Summary ──────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"Total companies : {n_total:,}")
    print(f"With description: {n_desc:,}  ({coverage:.1f}% coverage)")
    print(f"Without         : {n_total - n_desc:,}")
    print(f"\nCoverage by industry:")
    grp = df.groupby("Industry").apply(
        lambda g: pd.Series({
            "n": len(g),
            "with_desc": (g["description"].str.len() > 50).sum(),
        })
    )
    grp["pct"] = (grp["with_desc"] / grp["n"] * 100).round(0)
    grp = grp.sort_values("pct")
    print(grp[["n", "with_desc", "pct"]].to_string())

    # Financial metric coverage
    for col, label in [("rev_growth", "Rev growth"), ("ebitda_margin", "EBITDA margin"),
                       ("gross_margin", "Gross margin"), ("fcf_margin", "FCF margin")]:
        if col in df.columns:
            n = df[col].notna().sum()
            print(f"  {label:15s}: {n:,}/{n_total:,} ({n/n_total*100:.0f}%)")

    print(f"\nSample descriptions:")
    sample = df[df["description"].str.len() > 100].sample(min(5, n_desc), random_state=42)
    for _, r in sample.iterrows():
        print(f"\n  {r['Company']} ({r['Country']}) — match score {r['match_score']:.2f}")
        print(f"  {r['description'][:180]}…")


def main():
    print("=" * 60)
    print("Forbes Global 2000 — Description Enrichment")
    print("=" * 60)

    # Resume mode: if a prior enriched pkl exists, only re-fetch companies
    # that are still missing a description (handles mid-run auth failures).
    if OUT_PKL.exists():
        print(f"\nResuming from existing pickle: {OUT_PKL}")
        with open(OUT_PKL, "rb") as f:
            prior = pickle.load(f)
        df = prior["df"]
        n_have = (df.get("description", pd.Series()).str.len() > 50).sum()
        n_missing = (df.get("description", pd.Series()).str.len() <= 50).sum()
        print(f"  Already have: {n_have:,} descriptions")
        print(f"  Missing:      {n_missing:,} — will re-fetch these")
        if n_missing == 0:
            print("  Nothing to do — all companies already have descriptions.")
            save(df)
            return
        # Only enrich rows still missing a description
        need_idx = df[df.get("description", pd.Series()).str.len() <= 50].index
        partial  = pd.read_excel(INPUT).dropna(subset=["Company"])
        # keep only rows at positions matching need_idx
        partial  = partial.loc[partial.index.isin(need_idx) | (partial.index < len(df))]
        # Simpler: re-fetch using the original Forbes rows for missing ones
        df_source = pd.read_excel(INPUT).dropna(subset=["Company"]).copy()
        to_fetch  = df_source.loc[need_idx].copy()
        print(f"\nFetching {len(to_fetch):,} missing descriptions …\n")
        new_results = enrich(to_fetch, max_workers=12)
        # Merge back: update all enriched columns for re-fetched rows
        ALL_ENRICHED_COLS = ["ticker", "yf_name", "description", "net_debt_B",
                             "match_score", "rev_growth", "ebitda_margin",
                             "gross_margin", "fcf_margin"]
        for col in ALL_ENRICHED_COLS:
            if col in new_results.columns:
                df.loc[need_idx, col] = new_results[col].values
        enriched = df
    else:
        print(f"\nLoading {INPUT} …")
        df = pd.read_excel(INPUT)
        df = df.dropna(subset=["Company"]).copy()
        print(f"Loaded {len(df):,} companies across {df['Industry'].nunique()} industries")
        enriched = enrich(df, max_workers=12)

    # Augment pass: fetch financial metrics for companies with tickers
    # that are still missing those fields. Uses Ticker.info directly
    # (no Search API) so it avoids the 401 crumb-expiry rate limit.
    enriched = augment_financials(enriched, max_workers=12)

    save(enriched)


if __name__ == "__main__":
    if "--fix-net-debt" in sys.argv:
        fix_net_debt_currency()
    else:
        main()
